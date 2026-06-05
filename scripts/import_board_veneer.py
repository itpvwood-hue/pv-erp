"""
import_board_veneer.py
----------------------
Reads  Board and veneer.xlsx from the VANACHAI Warehouse folder
and upserts every unique CODE into the ERP materials table.

Base Board columns (row 3+):
  0  ACC CODE | 1 Description | 2 CODE | 3 TYPE | 4 Thickness(mm) |
  5  Width(mm)| 6 Length(mm)  | 7 GLUE | 8 FSC  | 9 Unit          |
  10 Width(in)| 11 Length(in) | 12 Pack size

Veneer columns (row 2+):
  0  ACC CODE | 1 Description | 2 CODE | 3 Width(mm) | 4 Length(mm) |
  5  V-THI(mm)| 6 Species     | 7 Cut  | 8 Mat.      | 9 Grade      |
  10 FSC      | 11 Unit        | 12 F/B

Rules:
- Deduplicate by CODE; first occurrence wins for descriptive fields.
- If CODE already in DB: UPDATE descriptive/dim fields; KEEP existing
  current_stock, reorder_point, unit_cost, price, name_th.
- If CODE not in DB: INSERT with stock=0, reorder_point=0, price=0.
- Generate a clean English name from the structured fields.
"""

import sys, sqlite3, re
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

XLSX = Path(r'C:\Users\PV_Natthapat\Desktop\VANACHAI\PV wood\Warehouse\Board and veneer.xlsx')
DB   = Path(r'C:\Users\PV_Natthapat\Desktop\Calude ERP\erp.db')

try:
    import openpyxl
except ImportError:
    print('Installing openpyxl…')
    import subprocess; subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

wb = openpyxl.load_workbook(XLSX, data_only=True)

# ── helpers ───────────────────────────────────────────────────────────────────
def clean(v):
    if v is None: return ''
    return str(v).strip().replace('\n', ' ').replace('  ', ' ')

def flt(v):
    try: return float(v) if v is not None else None
    except: return None

def fmt_num(v):
    if v is None: return ''
    f = float(v)
    return str(int(f)) if f == int(f) else str(f)

def board_name(typ, glue, thick, width, length):
    parts = []
    bt = clean(typ) if typ and typ != '-' else ''
    gt = clean(glue) if glue and glue != '-' else ''
    if bt: parts.append(bt)
    if gt and gt.upper() not in bt.upper():   # avoid "MDF FS E1 E1"
        parts.append(gt)
    t = flt(thick); w = flt(width); l = flt(length)
    if t: parts.append(fmt_num(t) + 'mm')
    if w and l: parts.append(fmt_num(w) + 'x' + fmt_num(l) + 'mm')
    return ' '.join(parts) if parts else 'Base Board'

def veneer_name(species, cut, thick, width, length, grade, match):
    parts = []
    if species: parts.append(clean(species))
    if cut:     parts.append(clean(cut))
    t = flt(thick); w = flt(width); l = flt(length)
    if t: parts.append(fmt_num(t) + 'mm')
    if w and l: parts.append(fmt_num(w) + 'x' + fmt_num(l) + 'mm')
    if grade and grade not in ('', '-'): parts.append(clean(grade))
    if match and match not in ('', '-'): parts.append(clean(match))
    return ' '.join(parts) if parts else 'Veneer'

# ── read sheets ───────────────────────────────────────────────────────────────
ws_board  = wb['Base Board Material ']
ws_veneer = wb['Veneer Material']

boards = {}   # code → dict
for row in ws_board.iter_rows(min_row=3, values_only=True):
    code = clean(row[2])
    if not code or code in boards:
        continue
    typ   = clean(row[3])
    thick = flt(row[4])
    width = flt(row[5])
    leng  = flt(row[6])
    glue  = clean(row[7])
    fsc   = clean(row[8])
    unit  = clean(row[9]) or 'sheet'
    acc   = clean(row[0])
    name  = board_name(typ, glue, thick, width, leng)  # always spec-generated
    boards[code] = dict(
        code=code, name=name, type='core_board', unit=unit,
        thickness_mm=thick, width_mm=width, length_mm=leng,
        acc_code=acc, board_type=typ if typ != '-' else '',
        glue_type=glue if glue != '-' else '',
        fsc=fsc if fsc != '-' else '',
    )

veneers = {}  # code → dict
for row in ws_veneer.iter_rows(min_row=2, values_only=True):
    code = clean(row[2])
    if not code or code in veneers:
        continue
    width   = flt(row[3])
    leng    = flt(row[4])
    thick   = flt(row[5])
    species = clean(row[6])
    cut     = clean(row[7])
    match   = clean(row[8])
    grade   = clean(row[9])
    fsc     = clean(row[10])
    unit    = clean(row[11]) or 'sheet'
    fb      = clean(row[12])
    acc     = clean(row[0])
    name    = veneer_name(species, cut, thick, width, leng, grade, match)
    veneers[code] = dict(
        code=code, name=name, type='veneer_sheet', unit=unit,
        thickness_mm=thick, width_mm=width, length_mm=leng,
        acc_code=acc, fsc=fsc if fsc != '-' else '',
        species=species, cut_type=cut, matching=match,
        grade=grade if grade not in ('', '-') else '',
        face_back=fb,
    )

print(f'Parsed: {len(boards)} base boards, {len(veneers)} veneers')

# ── upsert into DB ────────────────────────────────────────────────────────────
conn = sqlite3.connect(DB)
conn.row_factory = sqlite3.Row
conn.execute('PRAGMA journal_mode=WAL')

# Index existing materials by code
existing = {r['code']: r['id'] for r in conn.execute(
    "SELECT id, code FROM materials WHERE code != '' AND code IS NOT NULL"
).fetchall()}

ins_b = upd_b = ins_v = upd_v = 0

def upsert(mat, is_board):
    global ins_b, upd_b, ins_v, upd_v
    code = mat['code']
    mid  = existing.get(code)

    if mid:
        # UPDATE — only overwrite descriptive/dimension fields, preserve financials
        conn.execute("""
            UPDATE materials SET
              name=?, type=?, unit=?,
              thickness_mm=?, width_mm=?, length_mm=?,
              acc_code=?, fsc=?,
              board_type=?, glue_type=?,
              species=?, cut_type=?, matching=?, grade=?, face_back=?,
              is_active=1
            WHERE id=?
        """, (
            mat['name'], mat['type'], mat['unit'],
            mat.get('thickness_mm'), mat.get('width_mm'), mat.get('length_mm'),
            mat.get('acc_code',''), mat.get('fsc',''),
            mat.get('board_type',''), mat.get('glue_type',''),
            mat.get('species',''), mat.get('cut_type',''),
            mat.get('matching',''), mat.get('grade',''), mat.get('face_back',''),
            mid
        ))
        if is_board: upd_b += 1
        else: upd_v += 1
    else:
        # INSERT new
        cur = conn.execute("""
            INSERT INTO materials
              (code, name, type, unit,
               thickness_mm, width_mm, length_mm,
               current_stock, reorder_point, unit_cost, price,
               acc_code, fsc,
               board_type, glue_type,
               species, cut_type, matching, grade, face_back,
               is_active)
            VALUES (?,?,?,?, ?,?,?, 0,0,0,0, ?,?, ?,?, ?,?,?,?,?, 1)
        """, (
            mat['code'], mat['name'], mat['type'], mat['unit'],
            mat.get('thickness_mm'), mat.get('width_mm'), mat.get('length_mm'),
            mat.get('acc_code',''), mat.get('fsc',''),
            mat.get('board_type',''), mat.get('glue_type',''),
            mat.get('species',''), mat.get('cut_type',''),
            mat.get('matching',''), mat.get('grade',''), mat.get('face_back',''),
        ))
        existing[code] = cur.lastrowid
        if is_board: ins_b += 1
        else: ins_v += 1

for mat in boards.values():
    upsert(mat, is_board=True)

for mat in veneers.values():
    upsert(mat, is_board=False)

conn.commit()
conn.close()

print(f'Base boards : {ins_b} inserted, {upd_b} updated')
print(f'Veneers     : {ins_v} inserted, {upd_v} updated')
print(f'Total materials in import: {ins_b+upd_b+ins_v+upd_v}')
print('Done.')
